"""Create an Excel/WPS-friendly annotation workbook from a review CSV.

The workbook puts the human annotation fields and the evidence/QA context first,
adds dropdowns for label fields, freezes the header row, and wraps long text.

Usage:
    python -m scripts.create_review_workbook \
        --input outputs/runs/no_vote_500_newdef_qwen3_32b/human_review_100_zh.csv \
        --output outputs/runs/no_vote_500_newdef_qwen3_32b/human_review_100_zh.xlsx
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


FRONT_COLUMNS = [
    "sample_id",
    "annotation_priority",
    "suggested_difficulty_label",
    "full_context_qa",
    "required_evidence_ids",
    "evidence_context_zh",
    "qa_zh",
    "human_valid",
    "human_difficulty_label",
    "human_answer_directly_found",
    "human_evidence_ids",
    "human_notes",
    "full_context_numbered",
]
BACK_COLUMNS = ["evidence_context", "qa"]

COLUMN_WIDTHS = {
    "sample_id": 10,
    "annotation_priority": 14,
    "suggested_difficulty_label": 18,
    "full_context_qa": 76,
    "required_evidence_ids": 18,
    "evidence_context": 52,
    "evidence_context_zh": 52,
    "qa": 42,
    "qa_zh": 42,
    "human_valid": 14,
    "human_difficulty_label": 20,
    "human_answer_directly_found": 24,
    "human_evidence_ids": 20,
    "human_notes": 32,
    "full_context_numbered": 72,
}

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HUMAN_FILL = PatternFill("solid", fgColor="FFF2CC")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create xlsx review workbook with dropdown annotations.",
    )
    parser.add_argument("--input", required=True, help="Input review CSV.")
    parser.add_argument("--output", required=True, help="Output xlsx path.")
    return parser.parse_args()


def read_csv(path: Path) -> tuple[list[str], list[dict]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def ordered_columns(fieldnames: list[str]) -> list[str]:
    front = [c for c in FRONT_COLUMNS if c in fieldnames]
    back = [c for c in BACK_COLUMNS if c in fieldnames]
    rest = [c for c in fieldnames if c not in front and c not in back]
    return front + rest + back


def add_dropdown(ws, fieldnames: list[str], name: str, formula: str, max_row: int) -> None:
    if name not in fieldnames or max_row < 2:
        return
    col_idx = fieldnames.index(name) + 1
    col_letter = get_column_letter(col_idx)
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Please select a value from the dropdown list."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Select a value."
    validation.promptTitle = name
    ws.add_data_validation(validation)
    validation.add(f"{col_letter}2:{col_letter}{max_row}")


def style_sheet(ws, fieldnames: list[str], max_row: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HUMAN_FILL if str(cell.value).startswith("human_") else HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for idx, name in enumerate(fieldnames, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(name, 22)
        if name.startswith("human_"):
            for row in range(2, max_row + 1):
                ws.cell(row=row, column=idx).fill = HUMAN_FILL

    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 110


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    original_fields, rows = read_csv(input_path)
    fields = ordered_columns(original_fields)

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])

    max_row = len(rows) + 1
    add_dropdown(ws, fields, "human_valid", '"yes,no"', max_row)
    add_dropdown(ws, fields, "human_difficulty_label", '"Easy,Medium,Hard"', max_row)
    add_dropdown(ws, fields, "human_answer_directly_found", '"yes,no"', max_row)

    style_sheet(ws, fields, max_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved workbook: {output_path}")
    print(f"Rows: {len(rows)}")
    print("Front columns:")
    for field in fields[:len(FRONT_COLUMNS)]:
        print(f"  {field}")


if __name__ == "__main__":
    main()
